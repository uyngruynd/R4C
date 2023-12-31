import io
from datetime import datetime, timedelta

from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook

from robots.models import Robot


def get_date(text: str) -> datetime:
    date = datetime.strptime(text, '%Y-%m-%d %H:%M:%S')
    tz = timezone.get_current_timezone()
    tz_datetime = timezone.make_aware(date, tz, True)
    return tz_datetime


def get_serial(model, version):
    return f'{model}-{version}'


class RobotsReport:

    def __init__(self, start_date=None, end_date=None):
        if end_date:
            self.end_date = get_date(end_date)
        else:
            self.end_date = timezone.now()

        if start_date:
            self.start_date = get_date(start_date)
        else:
            self.start_date = self.end_date - timedelta(days=7)

    # noinspection PyMethodMayBeStatic
    def _add_row(self, worksheet, row_idx, data):
        for col_idx, value in enumerate(data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    def _get_report_stream(self):
        workbook = Workbook()
        models = Robot.objects.values_list('model', flat=True).distinct()

        for sheet_idx, model in enumerate(models):
            worksheet = workbook.create_sheet(title=model, index=sheet_idx)
            self._add_row(worksheet, 1,
                          ['Модель', 'Версия', 'Количество за неделю'])

            versions = Robot.objects.values('version').filter(
                created__range=[self.start_date, self.end_date],
                model=model).annotate(count=Count('version')).order_by(
                'version')

            for ver_idx, ver in enumerate(versions, start=2):
                self._add_row(worksheet, ver_idx,
                              [model, ver.get('version'), ver.get('count')])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def execute(self):
        """
        Возвращаю двоичные данные, содержащие сформированный отчет директору.
        """
        filename = f'RobotsReport-{datetime.now().strftime("%Y-%m-%d")}'

        response = HttpResponse(content=self._get_report_stream(),
                                content_type='application/ms-excel', )
        response[
            'Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response
